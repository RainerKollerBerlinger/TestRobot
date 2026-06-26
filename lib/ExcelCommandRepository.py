import sys
import os
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'PythonFunctions'))

from Command_set import command_set
from Command_read import command_read
from Read_Expected import read_expected

# ── RMU protocol constants — change here when targeting a different controller ─
COMMAND_ID   = '99'
MSG_TYPE_SET = '@'
MSG_TYPE_READ = '?'
DEST_ADDR    = 0x30   # 48 decimal
# ──────────────────────────────────────────────────────────────────────────────


class ExcelCommandRepository:
    """
    Loads the command repository from an Excel file.
    To swap to SharePoint later: implement SharePointCommandRepository
    with the same load_commands() / get_step() interface and change
    the Library line in the robot Settings section.
    """
    ROBOT_LIBRARY_SCOPE = 'SUITE'

    def load_commands(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self._commands = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if not data.get("BlockCmdID"):
                continue
            key = (data["BlockCmdID"], data["LimitType"])
            self._commands[key] = self._calculate_commands(data)

    def get_step(self, block_cmd_id, limit_type):
        key = (block_cmd_id, limit_type)
        if key not in self._commands:
            raise KeyError(f"No command found for BlockCmdID='{block_cmd_id}' LimitType='{limit_type}'")
        return self._commands[key]

    def _calculate_commands(self, data):
        table      = int(data["Table"])
        block      = int(data["Block"])
        offset     = int(data["ByteOffset"])
        bytes_val  = int(data["Bytes"])
        meas_value = data["MeasValue"]
        format_txt = data["BlockCmdID:FormatTXT"]
        divisor    = float(data["BlockCmdID:DivisorTXT"])
        set_expected = str(data["SetExpected"])

        set_command       = command_set(MSG_TYPE_SET, COMMAND_ID, DEST_ADDR,
                                        table, block, offset, bytes_val, meas_value,
                                        format_txt, divisor)
        read_command      = command_read(MSG_TYPE_READ, COMMAND_ID, DEST_ADDR,
                                         table, block, offset, bytes_val)
        read_expected_val = read_expected(MSG_TYPE_READ, COMMAND_ID, DEST_ADDR,
                                          table, block, offset, bytes_val, meas_value,
                                          format_txt, divisor)

        return {
            "BlockCmdID":   data["BlockCmdID"],
            "LimitType":    data["LimitType"],
            "TeleField":    data["TeleField"],
            "CloudField":   data["CloudField"],
            "SetCommand":   set_command,
            "SetExpected":  set_expected,
            "ReadCommand":  read_command,
            "ReadExpected": read_expected_val,
            "TeleExpected": str(data["TeleExpected"]),
            "CloudExpected":str(data["CloudExpected"]),
        }
